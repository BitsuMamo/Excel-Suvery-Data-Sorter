from openpyxl import Workbook, load_workbook
from os import path, remove

class SortData():
    """
    Class used to sort the excel data for engineering survey

    Attributes
    ----------
    unsorted_path (str):
        The path to the file to be sorted.
    sorted_path (str):
        The path to the file that is sorted.
    col_num (str):
        Column letter of number.
    col_x (str):
        Column letter of x.
    col_y (str):
        Column letter of y.
    col_ele (str):
        Column letter of elevation.
    col_cat (str):
        Column letter of category.
    col_con_1 (str):
        Column letter of x,y,ele to be concatenated.
    col_con_2 (str):
        Column letter of all concatenation.
    categories (list):
        All the uique categories is the file to be sorted.
    sizes (list):
        All the number of rows in the different sheet from the sorted data.
    unsorted_row_begin (int):
        The first row with data in the file to be sorted.
    unsorted_row_end (int):
        The last row with data in the file to be sorted.
    unsorted_work_sheet:
        The worksheet with data to be sorted.

    Methods
    -------
    get_categories():
        Gets the categories from the worksheet to be sorted.
    sort_data():
        Sorts the data from the unsorted file to a new file.
    get_row_sizes():
        Gets the row sizes of each sheet in sorted worksheet.
    concatenate():
        Concatenated all the data.
    concatenate_x_y_ele(work_sheet, row:int):
        concatenates the x, y and elevation of the data.
    concatenate_all(work_sheet, row:int):
        concatenates the x, y and elevation and the catagory, number of the data.
    run():
        Runs the class.

    """

    def __init__(self, unsorted_path, col_num, col_x, col_y, col_ele, col_cat, col_con_1, col_con_2) -> None:
        '''
        Constructor
            Parameters:
                unsorted_path (str): the path to the file to be sorted
                col_num (str): column letter of number
                col_x (str): column letter of x
                col_y (str): column letter of y
                col_ele (str): column letter of elevation
                col_cat (str): column letter of category
                col_con_1 (str): column letter of x,y,ele to be concatenated
                col_con_2 (str): column letter of all concatenation
            Return: None
        '''
        self.unsorted_path: str = unsorted_path
        self.col_num: str = col_num
        self.col_x: str = col_x
        self.col_y: str = col_y
        self.col_ele: str = col_ele
        self.col_cat: str = col_cat
        self.col_con_1: str = col_con_1
        self.col_con_2: str = col_con_2
        self.categories: list = []
        self.sizes:list = []
        if path.dirname(self.unsorted_path):
            self.sorted_path: str = str(path.dirname(self.unsorted_path)) + "/sorted_" + str(path.basename(self.unsorted_path))
        else:
            self.sorted_path: str = "sorted_" + str(path.basename(self.unsorted_path))

        if path.exists(self.unsorted_path):
            print("Loading unsorted workbook.\n")
            self.unsorted_work_book: Workbook = load_workbook(self.unsorted_path)
            self.unsorted_work_sheet  = self.unsorted_work_book.active
        else:
            print("File not found!\n")
            exit()

        if not(path.exists(self.sorted_path)):
            print("Creating sorted workbook.\n")
            self.sorted_work_book: Workbook = Workbook()
            self.sorted_work_book.save(self.sorted_path)
            self.sorted_work_book.close()
        else:
            print("Removing previous sorted workbook\n")
            remove(self.sorted_path)
            print("Creating sorted workbook.")
            self.sorted_work_book = Workbook()
            self.sorted_work_book.save(self.sorted_path)
            self.sorted_work_book.close()

        self.unsorted_row_being: int = self.unsorted_work_sheet.min_row
        self.unsorted_row_end: int = self.unsorted_work_sheet.max_row + 1

    def get_categories(self) -> None:

        '''
        Returns a list of all the categories found in the unsorted workbook

            Parameters:
                work_sheet (WorkSheet): The worksheet to be sorted
            Returns:
                categories (list): A list of all uique categories
        '''

        for row in range(self.unsorted_row_being, self.unsorted_row_end):
            cell = self.col_cat + str(row)

            if not(str(self.unsorted_work_sheet[cell].value).upper() in self.categories):
                self.categories.append(str(self.unsorted_work_sheet[cell].value).upper())

        if "NONE" in self.categories:
            self.categories.remove("NONE")

        print(f"Categories : {self.categories}\nNumber of Categories: {len(self.categories)}\n")

    def sort_data(self) -> None:

        '''
        Creates a sorted work book with all the sheets created and all the data populated.

            Parameters: None
            Returns: None
        '''

        print("Sorting Data.\n")

        for category in self.categories:
            self.sorted_work_book.create_sheet(category)
            sorted_work_sheet = self.sorted_work_book[category]

            for row in range(self.unsorted_row_being, self.unsorted_row_end):
                cell = self.col_cat + str(row)

                if str(self.unsorted_work_sheet[cell].value).upper() == category:
                    data = [self.unsorted_work_sheet[self.col_num + str(row)].value,
                            self.unsorted_work_sheet[self.col_x + str(row)].value,
                            self.unsorted_work_sheet[self.col_y + str(row)].value,
                            self.unsorted_work_sheet[self.col_ele + str(row)].value,
                            self.unsorted_work_sheet[self.col_cat + str(row)].value]

                    sorted_work_sheet.append(data)

        self.sorted_work_book.save(self.sorted_path)
        self.sorted_work_book.close()

    def get_row_sizes(self) -> None:

        '''
        Returns a list of all rows occupied in each sheet of a workbook

            Parameters: None
            Returns:
                sizes (list): all the rows in each sheet
        '''

        self.sorted_work_book = load_workbook(self.sorted_path)

        for category in self.categories:
            work_sheet = self.sorted_work_book[category]
            self.sizes.append(work_sheet.max_row + 1)

        self.sorted_work_book.close()

    def concatenate(self) -> None:

        '''
        Adds the concatenated string to the worksheet

            Parameters: None
            Returns: None
        '''

        print("Concatenating sorted data.\n")
        self.get_row_sizes()
        self.sorted_work_book = load_workbook(self.sorted_path)
        for index, category in enumerate(self.categories):
            sorted_work_sheet = self.sorted_work_book[category]

            for row in range(1, self.sizes[index]):
                sorted_work_sheet[self.col_con_1 + str(row)] = self.concatenate_x_y_ele(sorted_work_sheet, row)
                sorted_work_sheet[self.col_con_2 + str(row)] = self.concatenate_all(sorted_work_sheet, row)

        self.sorted_work_book.save(self.sorted_path)
        self.sorted_work_book.close()

    def concatenate_x_y_ele(self, work_sheet, row) -> str:


        '''
        Returns a string with the x, y and z concatenated with a comma

            Parameters:
                work_sheet (WorkSheet): The worksheet to retrive the data from
                row (int): The row to be concatenated
            Returns:
                (str): The concatenated value
        '''
        row = str(row)
        return str(work_sheet[self.col_x + row].value) + ',' + str(work_sheet[self.col_y + row].value) + \
                ',' + str(work_sheet[self.col_ele + row].value)

    def concatenate_all(self, work_sheet, row) -> str:

        '''
        Returns a string with the x, y, z, num, and category concatenated with a comma

            Parameters:
                work_sheet (WorkSheet): The worksheet to retrive the data from
                row (int): The row to be concatenated
            Returns:
                (str): The concatenated value
        '''

        row = str(row)
        return str(work_sheet[self.col_num + row].value) + ',' + str(work_sheet[self.col_x + row].value) + ',' \
                + str(work_sheet[self.col_y + row].value) + ',' + str(work_sheet[self.col_ele + row].value) \
                + ',' + str(work_sheet[self.col_cat + row].value)

    def run(self) -> None:

        '''
        Runs the class to sort data
            Parameters: None
            Returns: None
        '''

        self.get_categories()
        self.sort_data()
        self.concatenate()

        self.unsorted_work_book.close()
        self.sorted_work_book.close()

        # Debuggin puropses
        print(f"Path of Unsorted: {self.unsorted_path}\nPath of Sorted: {self.sorted_path}\n")


if __name__ == "__main__":
    #sort = SortData(input("Enter Path: "), input("Column of Num: "), input("Column X: "), input("Column Y: "), input("Column Elevation: "), input("Column Type"), input("Column For X, Y, ELE Concat: "), input("Column For All Concat: "))
    print(SortData.__doc__)
    #sort.run()
